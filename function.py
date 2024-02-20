# функции для получения отчета
# Разгрузка основного документа

import os
import openpyxl
import requests
from environs import Env
from datetime import datetime
from dateutil.relativedelta import relativedelta


# Данные о путях и токены для доступа к API берем из .env файла
dotenv_path = os.path.join(os.path.dirname(__file__), '.env')
if os.path.exists(dotenv_path):
    env = Env()
    env.read_env(dotenv_path)
else:
    exit('Отсутствует файл конфигурации.')

TOKEN_WB = env("TOKEN_WB")
SHEET_URL = env("SHEET_URL")  # Документ Google
ARTICLE_DICT = env("ARTICLE_DICT")  # Словарь сопоставления


def old_date(mon):
    today = datetime.now()
    new_date = today - relativedelta(months=mon)  # используем оригинальную дату и вычитаем количество месяцев
    return new_date.strftime('%Y-%m-%d')


def get_google_document():
    """Скачиваем гугл документ и возвращаем его в виде словаря:
    {'Название колонки': [Список значений колонки]}

    Документ сохраняется с именем google_sheet_data.xlsx.
    Листы, которые читаются из документа перечислены в файле sheets_google.txt.
    Названия колонок берутся из первой строки первого листа.
    Выбирается первые 10 колонок.
    """
    # Получаем данные из Google Sheets
    response = requests.get(SHEET_URL + '/export?format=xlsx')
    if response.status_code != 200:
        exit('Ошибка чтения Google документа')

    with open('google_sheet_data.xlsx', 'wb') as file:
        file.write(response.content)
    response.close()

    # Открываем Excel-документ
    wb = openpyxl.load_workbook('google_sheet_data.xlsx', data_only=True)
    google_sheet_data = {}
    columns = dict()  # {name: num}
    # Подготовка списка листов Google документа из которых будут взяты товары
    if not os.path.exists('sheets_google.txt'):
        exit('Отсутствует файл конфигурации листов Googl документа.')

    with (open('sheets_google.txt', 'r', encoding='utf-8') as file):
        for line in file.readlines():
            value = line.strip()
            sheet = wb[value]

            # Проходимся по каждой ячейке в первой строке до 10 (заголовки столбцов)
            if not google_sheet_data:
                head = True
                for col in range(1, 11):
                    field_name = sheet.cell(row=1, column=col).value
                    google_sheet_data[field_name] = []
                    columns[field_name] = col

            # Проходимся по строкам, начиная со второй (в первой шапка таблицы)
            for row in range(2, sheet.max_row + 1):
                # Проходимся по столбцам
                if sheet.cell(row=row, column=columns['артикул']).value and sheet.cell(
                        row=row, column=columns['Штрихкод']).value:
                    # Записи без этих данных будут пропущены
                    for col, i in columns.items():
                        cell_value = sheet.cell(row=row, column=i).value
                        google_sheet_data[col].append(cell_value)

    return google_sheet_data


def get_article_dict():
    """Загрузка словаря сопоставления артикулов с Yandex диска"""
    response = requests.get(ARTICLE_DICT)
    if response.status_code != 200:
        exit('Ошибка чтения словаря Артикулов')

    with open('article_dict.xlsx', 'wb') as file:
        file.write(response.content)
    response.close()

    # Открываем Excel-документ
    wb = openpyxl.load_workbook('article_dict.xlsx', data_only=True)
    sheet = wb['Лист1']
    article_dict = dict()
    for row in range(2, sheet.max_row + 1):
        article_dict[str(sheet.cell(row=row, column=2).value)] = str(sheet.cell(row=row, column=1).value)

    return article_dict

