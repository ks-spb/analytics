# Отчет использует Google документ и API Wb

import string
from openpyxl import Workbook, styles
from openpyxl.styles import Font

from calculator import Calculator


calculator = Calculator()  # Создаем объект со всеми исходными данными и методами расчета полей

# Регистрация методов для рассчета полей

fields = {
    'Наименование товара': calculator.product_name,
    'Артикул продавца': calculator.product_article,
    'Артикул WB': calculator.product_article_wb,
    'Штрихкод': calculator.barcode,
    'Статус': calculator.status,
    'Заказано': calculator.ordered,
    'Сумма заказов с комиссией WB': calculator.ordered_sum,
    'Выкупили': calculator.purchased,
    'К перечислению': calculator.purchased_sum,
    'Остаток': calculator.leftover,
    'Стоимость 1 шт.': calculator.price_one_item,
    'Дней продаж': calculator.sales_days,
    'Скорость продаж': calculator.sales_speed,
    'Остатка хватит, дней': calculator.days_left,
    'Дней до отгрузки': calculator.days_before_delivery,
    'На сколько дней поставка': calculator.days_left_to_delivery,
    'Старый вариант поставки': calculator.old_delivery,
    'Поставка по стоимости (фильтр)': calculator.delivery_filter,
    'Код 1С': calculator.code_1c,
    'Спецификация': calculator.specification,
    'Скидка': calculator.discount,
    'Стоимость до скидки': calculator.price_before_discount,
    'Цена после скидки': calculator.price_after_discount,
    'Разница': calculator.difference,
}

# Ширина столбцов по номерам
width_column = [80, 15, 15, 15] + [12]*30

# Создаем новый документ
wb = Workbook()

# Получаем активный лист
ws = wb.active

# Добавляем заголовки полей в первую строку
ws.append([i for i in fields.keys()])

# Устанавливаем высоту первой строки
ws.row_dimensions[1].height = 60

# Добавляем данные.
# Создаем список свойств (артикул WB, Наименование товара) не проданных товаров
no_sales_products = []
while not calculator.get_next_article() is None:
    row = []
    for name_field, method in fields.items():
        try:
            row.append(method())
        except:
            pass

    if calculator.result['purchased'] == 0:
        # Этот товар не продавался в отчетный период, запоминаем его артикул
        no_sales_products.append(
            (calculator.result['status'],
             calculator.result['product_article'],
             calculator.result['product_name'],
             calculator.result['leftover'])
        )

    ws.append(row)

# Закрепление областей
ws.freeze_panes = 'D2'

# Устанавливаем перенос по словам для ячеек в диапазоне от 1-20 первой строки
for col in range(1, len(width_column)+1):
    cell = ws.cell(row=1, column=col)
    cell.alignment = styles.Alignment(wrap_text=True, horizontal='center', vertical='center')
    ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width_column[col-1]
    cell.font = Font(bold=True)


# Добавление листа со списком не проданных товаров
# ------------------------------------------------------------------
no_sales_sheet = wb.create_sheet(title="Не проданные товары")

# Добавляем заголовки полей в первую строку
no_sales_sheet.append(['Статус', 'Артикул', 'Наименование товара', 'Остаток'])

# Устанавливаем высоту первой строки
no_sales_sheet.row_dimensions[1].height = 30

# Ширина столбцов по номерам
width_column = [15, 15, 80, 15]

# Устанавливаем свойства столбцов
for col in range(1, 5):
    cell = no_sales_sheet.cell(row=1, column=col)
    cell.alignment = styles.Alignment(wrap_text=True, horizontal='center', vertical='center')
    no_sales_sheet.column_dimensions[no_sales_sheet.cell(row=1, column=col).column_letter].width = width_column[col-1]
    cell.font = Font(bold=True)

# Заполнение листа данными
for prop in no_sales_products:
        no_sales_sheet.append(prop)

# # Применение цветов
# red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
# # ws['A1':'A5'].apply(lambda x: setattr(x, 'fill', red_fill))
# for row in ws['A1:A5']:
#     for cell in row:
#         cell.fill = red_fill


# Сохраняем документ
old = calculator.old_date.replace('-', '.')
to = calculator.to_date.replace('-', '.')
wb.save(f'Report {old} - {to}.xlsx')

